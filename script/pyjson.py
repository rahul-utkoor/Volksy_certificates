import os, sys
import xlrd
import csv
import pandas as pd
from certificates import processCertificates 
import exceptions
import logging
import codecs
import openpyxl as px
from xlrd import open_workbook
from openpyxl import load_workbook
import json
from pprint import pprint
import numpy as np
import re
import urllib, json

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.FileHandler('Logs/scan.log')
logger.addHandler(handler)


def scan():	
	try:
		#gobal inputPath
		wb = load_workbook('Input/batch3/GDH .xlsx', read_only=True)
		ws = wb.get_sheet_by_name('Sheet1')
		use_col = 0
		data = []
		for r in ws.iter_rows() :
			temp = []
			for c in range (len(r)) :
				if type(r[c].value) == unicode :
					temp.append((r[c].value).encode('utf-8'))
				else :
					temp.append(r[c].value)

			data.append(temp)
			
		
		concat = []
		for i in data :
			temp = i[3] + i[8]
			temp = re.sub('\ ','',temp)
			temp = re.sub('\.','',temp)
			concat.append(temp)
		url="http://52.11.242.228:8080/api/public/getByConcatNames?names="+",".join(concat[1:])
		response = urllib.urlopen(url)
		data1 = json.loads(response.read())
		
		for i in range (len(data1["detailedCandidateData"])) :
			loct = data1["detailedCandidateData"][i]["location"]
			if not os.path.exists("Output/"+loct):
				os.makedirs("Output/"+loct)
			#editedFolder = processCertificates(folder, errFolder, notFound)
		
if __name__ == '__main__':
	inputPath = sys.argv[1]
	outputPath = sys.argv[2]
	print scan()
